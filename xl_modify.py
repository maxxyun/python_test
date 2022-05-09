#https://ybworld.tistory.com/11?category=936724

import os
import openpyxl

path = "D:\Git\python_test\xl_modify" #엑셀파일이 잇는 경로 설정
file_list = os.listdir(path) # path 폴더에 있는 파일들을 리스트로 받기
print(file_list)


num = int(input("몇개 바꿀지 정수로 입력 : "))
#동적변수 생성() : 바꿀 값이 여러개일경우 설정하기 위한 부분
for i in range(1, num+1):
    print(str(i)+"번째 설정:")
    globals()['cell_position'+str(i)] = str(input("바꿀 위치 : "))
    globals()['cell_value'+str(i)] = str(input("바꿀 문자열 : "))
    print("\n")


for file_name_raw in file_list:
    file_name = "D:\Git\python_test\xl_modify" + file_name_raw #실행 할 파일 경로 설정
    wb = openpyxl.load_workbook(filename=file_name) # workbook 객체 생성
    ws = wb.active # 워크북에서 활성화 된 시트 객체 설정
    #ws=wb['Sheet1']

    for i in range(1, num+1):
        print(file_name_raw +"의 "+str(i)+"번째 변경")
        cell_position = globals()['cell_position'+str(i)] #동적변수에 있는 값 저장(셀위치)
        cell_value = globals()['cell_value'+str(i)] # 동적변수에 있는 값 저장(셀값)
        ws[cell_position].value = cell_value # 정해진 셀위치에 셀값 입력
        #ws['A1'].value = "YB"
        print(cell_position+"의 값을("+cell_value+")로 입력합니다.")
    print("\n")
    wb.save(file_name_raw) #저장

